import openpyxl
from openpyxl.comments import Comment

workbook = openpyxl.load_workbook("students_attendance.xlsx")
sheet=workbook["Sheet1"]

max_row = sheet.max_row
max_column = sheet.max_column

#read cell
for i in range(1,max_row+1):
    attendance = sheet.cell(row=i, column=3).value
    if attendance == "Absent":
        name = sheet.cell(row=i,column=1).value
        email = sheet.cell(row=i,column=2).value
        print(name + " is absent")

#add value
sheet['A7'].value='Felicia'
sheet['B7'].value='Felicia@gmail.com'
sheet['C7'].value='Present'

#add comment
sheet['A7'].comment= Comment('Change text automatically','User')

#add a new element that count the number of non empty cell
#sheet['D7'] = '=COUNTA(A2:A50)'

#save the file
workbook.save("students_attendance_comment.xlsx")