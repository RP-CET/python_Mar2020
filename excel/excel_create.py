import openpyxl

workbook = openpyxl.Workbook()

#get the default sheet
sheet=workbook["Sheet"]

#create a list of tuples as data source
data = [
    [225.7,'Gone with the Wind','Victor Fleming'],
    [194.4, 'Star Wars', 'George Lucas'],
    [161.0, 'ET: The Extraterrestrial', 'Steven Spielberg']
]

#update value into cell
row = 1
for (admissions,name, director) in data:
    sheet['A'+str(row)].value = admissions
    sheet['B{}'.format(row)].value = name
    row = row + 1
    
#create a new sheet
sheet = workbook.create_sheet("Directors")

#print out added sheet name
print(workbook.sheetnames)

#update value into cell
for row, (admissions,name, director) in enumerate(data,1):
    sheet['A{}'.format(row)].value = director
    sheet['B{}'.format(row)].value = name

#save the spreadsheet
workbook.save("movies1.xlsx")