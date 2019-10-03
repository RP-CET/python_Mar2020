import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

workbook = openpyxl.Workbook()

# create a list of tuples as data source
data = [
  ['Name','Admission'],
  ['Gone with the Wind',225.7],
  ['Star Wars',161.0],
  ['ET: The Extraterrestrial',161.0]
  ]
   
sheet = workbook['Sheet']
for row in data:
  sheet.append(row)
  
#define the colors to use for styling
BLUE = "0033CC"
LIGHT_BLUE = "E6ECFF"
WHITE = "FFFFFF"

#define styling
header_font = Font(name="Tahoma", size=14, color=WHITE)
header_fill = PatternFill("solid", fgColor=BLUE)

# format header
for row in sheet["A1:B1"]:
  for cell in row:
    cell.font = header_font
    cell.fill = header_fill

#define styling
white_side = Side(border_style="thin", color=WHITE)
blue_side = Side(border_style="thin", color=BLUE)
alternate_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
border = Border(bottom=blue_side, left=white_side, right=white_side)

# format rows
for row_index, row in enumerate(sheet["A2:B5"]):
  for cell in row:
    cell.border = border
    if row_index %2 :
      cell.fill = alternate_fill

workbook.save("movie_format.xlsx")