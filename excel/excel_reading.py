import openpyxl

workbook = openpyxl.load_workbook("students_attendance.xlsx")
sheet=workbook["Sheet1"]

#You can use this to find the available sheets
# print(workbook.sheetnames)

max_row = sheet.max_row
max_column = sheet.max_column

#loop through every row
for i in range(1,max_row+1):
    
    #read cell
    attendance = sheet.cell(row=i, column=3).value
    
    #check attendance
    if attendance == "Absent":
        name = sheet.cell(row=i,column=1).value
        email = sheet.cell(row=i,column=2).value
        print(name + " is absent")

